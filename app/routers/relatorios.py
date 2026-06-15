from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import pegar_banco
from app.models.animal import Animal
from app.models.producao import Producao, PrecoLeite
from app.auth import pegar_usuario_atual
from app.models.usuario import Usuario
from app.logger import logger_rel
from datetime import date, timedelta
from decimal import Decimal
import io

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

roteador = APIRouter(
    prefix="/relatorios",
    tags=["Relatórios"]
)

ANO_MINIMO = 2000
ANO_MAXIMO = 2100


def validar_periodo(ano: int, mes: int):
    hoje = date.today()
    if not ano:
        ano = hoje.year
    if not mes:
        mes = hoje.month
    if mes < 1 or mes > 12:
        raise HTTPException(status_code=400, detail="Mês deve ser entre 1 e 12.")
    if ano < ANO_MINIMO or ano > ANO_MAXIMO:
        raise HTTPException(status_code=400, detail=f"Ano deve ser entre {ANO_MINIMO} e {ANO_MAXIMO}.")
    return ano, mes


def calcular_periodo(ano: int, mes: int):
    inicio = date(ano, mes, 1)
    fim = date(ano, mes + 1, 1) - timedelta(days=1) if mes < 12 else date(ano + 1, 1, 1) - timedelta(days=1)
    return inicio, fim


def pegar_preco_vigente(usuario_id: int, data_ref: date, banco: Session):
    return banco.query(PrecoLeite).filter(
        PrecoLeite.usuario_id == usuario_id,
        PrecoLeite.vigente_a_partir <= data_ref
    ).order_by(PrecoLeite.vigente_a_partir.desc()).first()


NOMES_MESES = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
               "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]


# ─── PDF ─────────────────────────────────────────────────────────────────────

@roteador.get("/pdf/producao-mensal")
def exportar_pdf_producao_mensal(
    ano: int = None,
    mes: int = None,
    banco: Session = Depends(pegar_banco),
    usuario: Usuario = Depends(pegar_usuario_atual)
):
    ano, mes = validar_periodo(ano, mes)
    inicio, fim = calcular_periodo(ano, mes)

    try:
        preco = pegar_preco_vigente(usuario.id, fim, banco)
        preco_litro = preco.preco_litro if preco else Decimal("0")

        animais = banco.query(Animal).filter(
            Animal.usuario_id == usuario.id,
            Animal.status == "ativo",
            Animal.sexo == "F"
        ).all()

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)

        styles = getSampleStyleSheet()
        titulo_style = ParagraphStyle("titulo", parent=styles["Title"],
                                      fontSize=16, spaceAfter=6, alignment=TA_CENTER)
        subtitulo_style = ParagraphStyle("subtitulo", parent=styles["Normal"],
                                         fontSize=11, spaceAfter=12, alignment=TA_CENTER)
        normal_style = styles["Normal"]

        elementos = []
        elementos.append(Paragraph("Gestão de Gado Leiteiro", titulo_style))
        elementos.append(Paragraph(
            f"Relatório de Produção — {NOMES_MESES[mes-1]}/{ano}", subtitulo_style
        ))
        elementos.append(Paragraph(
            f"Produtor: {usuario.nome_completo or usuario.username}", normal_style
        ))
        elementos.append(Paragraph(
            f"Preço do leite: R$ {preco_litro:.2f}/litro", normal_style
        ))
        elementos.append(Spacer(1, 0.5*cm))

        dados_tabela = [["Animal", "Brinco", "Total Litros", "Aproveitado", "Descartado", "Valor (R$)"]]
        total_geral = Decimal("0")
        total_aproveitado = Decimal("0")
        total_descartado = Decimal("0")

        for animal in animais:
            producoes = banco.query(Producao).filter(
                Producao.animal_id == animal.id,
                Producao.data.between(inicio, fim)
            ).all()
            if not producoes:
                continue

            total = sum(p.quantidade_litros for p in producoes)
            aproveitado = sum(p.quantidade_litros for p in producoes if p.status == "aproveitado")
            descartado = total - aproveitado
            valor = aproveitado * preco_litro

            total_geral += total
            total_aproveitado += aproveitado
            total_descartado += descartado

            dados_tabela.append([
                animal.nome, animal.brinco,
                f"{total:.2f}L", f"{aproveitado:.2f}L",
                f"{descartado:.2f}L", f"R$ {valor:.2f}"
            ])

        dados_tabela.append([
            "TOTAL", "",
            f"{total_geral:.2f}L",
            f"{total_aproveitado:.2f}L",
            f"{total_descartado:.2f}L",
            f"R$ {total_aproveitado * preco_litro:.2f}"
        ])

        tabela = Table(dados_tabela, colWidths=[4*cm, 3*cm, 3*cm, 3*cm, 3*cm, 3*cm])
        tabela.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E7D32")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F1F8E9")]),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#C8E6C9")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))

        elementos.append(tabela)
        elementos.append(Spacer(1, 0.5*cm))
        elementos.append(Paragraph(
            f"Relatório gerado em: {date.today().strftime('%d/%m/%Y')}", normal_style
        ))

        doc.build(elementos)
        buffer.seek(0)

        logger_rel.info(f"PDF produção mensal gerado | {mes}/{ano} | usuário: {usuario.id}")

        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=producao_{mes:02d}_{ano}.pdf"}
        )
    except HTTPException:
        raise
    except Exception:
        logger_rel.error(f"Erro ao gerar PDF produção | {mes}/{ano} | usuário: {usuario.id}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erro ao gerar relatório PDF. Tente novamente."
        )


# ─── EXCEL ───────────────────────────────────────────────────────────────────

@roteador.get("/excel/producao-mensal")
def exportar_excel_producao_mensal(
    ano: int = None,
    mes: int = None,
    banco: Session = Depends(pegar_banco),
    usuario: Usuario = Depends(pegar_usuario_atual)
):
    ano, mes = validar_periodo(ano, mes)
    inicio, fim = calcular_periodo(ano, mes)

    try:
        preco = pegar_preco_vigente(usuario.id, fim, banco)
        preco_litro = preco.preco_litro if preco else Decimal("0")

        animais = banco.query(Animal).filter(
            Animal.usuario_id == usuario.id,
            Animal.status == "ativo",
            Animal.sexo == "F"
        ).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Produção {NOMES_MESES[mes-1]}"

        verde_escuro = PatternFill("solid", fgColor="2E7D32")
        verde_claro = PatternFill("solid", fgColor="C8E6C9")
        verde_linha = PatternFill("solid", fgColor="F1F8E9")
        fonte_branca = Font(color="FFFFFF", bold=True, size=11)
        fonte_bold = Font(bold=True, size=11)
        borda = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        centralizado = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A1:F1")
        ws["A1"] = f"Relatório de Produção — {NOMES_MESES[mes-1]}/{ano}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:F2")
        ws["A2"] = f"Produtor: {usuario.nome_completo or usuario.username} | Preço: R$ {preco_litro:.4f}/L"
        ws["A2"].alignment = Alignment(horizontal="center")

        cabecalhos = ["Animal", "Brinco", "Total Litros", "Aproveitado", "Descartado", "Valor (R$)"]
        for col, cab in enumerate(cabecalhos, 1):
            cell = ws.cell(row=4, column=col, value=cab)
            cell.fill = verde_escuro
            cell.font = fonte_branca
            cell.alignment = centralizado
            cell.border = borda

        linha = 5
        total_geral = Decimal("0")
        total_aproveitado = Decimal("0")
        total_descartado = Decimal("0")

        for i, animal in enumerate(animais):
            producoes = banco.query(Producao).filter(
                Producao.animal_id == animal.id,
                Producao.data.between(inicio, fim)
            ).all()
            if not producoes:
                continue

            total = sum(p.quantidade_litros for p in producoes)
            aproveitado = sum(p.quantidade_litros for p in producoes if p.status == "aproveitado")
            descartado = total - aproveitado
            valor = float(aproveitado * preco_litro)

            total_geral += total
            total_aproveitado += aproveitado
            total_descartado += descartado

            fill = verde_linha if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            valores = [animal.nome, animal.brinco, float(total),
                       float(aproveitado), float(descartado), valor]

            for col, val in enumerate(valores, 1):
                cell = ws.cell(row=linha, column=col, value=val)
                cell.fill = fill
                cell.alignment = centralizado
                cell.border = borda
            linha += 1

        totais = ["TOTAL", "", float(total_geral), float(total_aproveitado),
                  float(total_descartado), float(total_aproveitado * preco_litro)]
        for col, val in enumerate(totais, 1):
            cell = ws.cell(row=linha, column=col, value=val)
            cell.fill = verde_claro
            cell.font = fonte_bold
            cell.alignment = centralizado
            cell.border = borda

        larguras = [20, 12, 14, 14, 14, 14]
        for col, largura in enumerate(larguras, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = largura

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger_rel.info(f"Excel produção mensal gerado | {mes}/{ano} | usuário: {usuario.id}")

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=producao_{mes:02d}_{ano}.xlsx"}
        )
    except HTTPException:
        raise
    except Exception:
        logger_rel.error(f"Erro ao gerar Excel produção | {mes}/{ano} | usuário: {usuario.id}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erro ao gerar planilha Excel. Tente novamente."
        )


@roteador.get("/excel/rebanho")
def exportar_excel_rebanho(
    banco: Session = Depends(pegar_banco),
    usuario: Usuario = Depends(pegar_usuario_atual)
):
    try:
        animais = banco.query(Animal).filter(
            Animal.usuario_id == usuario.id
        ).order_by(Animal.nome).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rebanho"

        verde_escuro = PatternFill("solid", fgColor="2E7D32")
        verde_linha = PatternFill("solid", fgColor="F1F8E9")
        fonte_branca = Font(color="FFFFFF", bold=True, size=11)
        borda = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        centralizado = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A1:J1")
        ws["A1"] = f"Relatório de Rebanho — {date.today().strftime('%d/%m/%Y')}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        cabecalhos = ["Nome", "Brinco", "Raça", "Sexo", "Nascimento",
                      "Status", "Status Reprodutivo", "Prod. Diária (L)", "Peso (kg)", "Observação"]
        for col, cab in enumerate(cabecalhos, 1):
            cell = ws.cell(row=3, column=col, value=cab)
            cell.fill = verde_escuro
            cell.font = fonte_branca
            cell.alignment = centralizado
            cell.border = borda

        for i, animal in enumerate(animais):
            fill = verde_linha if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            valores = [
                animal.nome, animal.brinco, animal.raca or "",
                "Fêmea" if animal.sexo == "F" else "Macho",
                animal.nascimento.strftime("%d/%m/%Y") if animal.nascimento else "",
                animal.status, animal.status_reprodutivo,
                float(animal.producao_diaria_litros or 0),
                animal.peso_kg or "", animal.observacao or ""
            ]
            for col, val in enumerate(valores, 1):
                cell = ws.cell(row=i+4, column=col, value=val)
                cell.fill = fill
                cell.alignment = centralizado
                cell.border = borda

        larguras = [20, 12, 15, 8, 12, 12, 18, 14, 10, 25]
        for col, largura in enumerate(larguras, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = largura

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger_rel.info(f"Excel rebanho gerado | usuário: {usuario.id}")

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=rebanho_{date.today()}.xlsx"}
        )
    except HTTPException:
        raise
    except Exception:
        logger_rel.error(f"Erro ao gerar Excel rebanho | usuário: {usuario.id}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erro ao gerar planilha de rebanho. Tente novamente."
        )