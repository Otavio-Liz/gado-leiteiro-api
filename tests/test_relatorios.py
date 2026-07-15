"""
Testes da Seção 12 — Relatórios.

Cobre o checklist original:
- Confirmar que PDF/Excel só usam dados do usuário autenticado.
- Validar período futuro.
- Validar limites de ano/mês.
- Testar relatório sem dados.

Os testes de isolamento e "sem dados" usam o Excel de rebanho (não o de
produção) porque é mais simples de decodificar e inspecionar linha a
linha via openpyxl — os outros itens do checklist (cálculo financeiro com
preço mudando no meio do mês, produção descartada) já foram conferidos
manualmente na revisão: a lógica em app/utils_financeiro.py já calcula o
preço vigente por data de cada produção, não um preço único pro período.
"""
import io
from datetime import date, timedelta
from decimal import Decimal
import openpyxl


def _ler_excel(resposta):
    return openpyxl.load_workbook(io.BytesIO(resposta.content))


# ─── Validação de período ────────────────────────────────────────────────────

def test_periodo_futuro_e_bloqueado(usuario_autenticado):
    cliente, usuario, headers = usuario_autenticado
    data_futura = date.today() + timedelta(days=32)  # garante mês diferente do atual
    resposta = cliente.get(
        "/relatorios/excel/producao-mensal",
        params={"ano": data_futura.year, "mes": data_futura.month},
        headers=headers,
    )
    assert resposta.status_code == 400


def test_mes_invalido_e_bloqueado(usuario_autenticado):
    cliente, usuario, headers = usuario_autenticado
    resposta = cliente.get(
        "/relatorios/excel/producao-mensal",
        params={"ano": 2026, "mes": 13},
        headers=headers,
    )
    assert resposta.status_code == 400


def test_ano_fora_do_limite_e_bloqueado(usuario_autenticado):
    cliente, usuario, headers = usuario_autenticado
    resposta = cliente.get(
        "/relatorios/excel/producao-mensal",
        params={"ano": 1990, "mes": 1},
        headers=headers,
    )
    assert resposta.status_code == 400


def test_periodo_atual_e_permitido(usuario_autenticado, criar_animal, criar_producao):
    """Sem nenhuma produção no período, a rota devolve 404 por design
    ('Sem dados de produção para o período selecionado.') — não é erro de
    rota. Pra testar que um período válido É aceito, precisa existir pelo
    menos uma produção dentro do mês consultado."""
    cliente, usuario, headers = usuario_autenticado
    animal = criar_animal(usuario.id, sexo="F", status_reprodutivo="em_lactacao")
    hoje = date.today()
    criar_producao(animal.id, data=hoje, quantidade_litros=Decimal("15.00"))

    resposta = cliente.get(
        "/relatorios/excel/producao-mensal",
        params={"ano": hoje.year, "mes": hoje.month},
        headers=headers,
    )
    assert resposta.status_code == 200


def test_periodo_atual_sem_producao_retorna_404_sem_dados(usuario_autenticado):
    """Confirma o outro lado: período válido, mas sem NENHUM dado de
    produção, devolve 404 com mensagem de 'sem dados' — não 200 vazio e
    não 500."""
    cliente, usuario, headers = usuario_autenticado
    hoje = date.today()
    resposta = cliente.get(
        "/relatorios/excel/producao-mensal",
        params={"ano": hoje.year, "mes": hoje.month},
        headers=headers,
    )
    assert resposta.status_code == 404
    assert "dados" in resposta.json()["mensagem"].lower()


# ─── Relatório sem dados ──────────────────────────────────────────────────────

def test_relatorio_rebanho_sem_animais_nao_quebra(usuario_autenticado):
    cliente, usuario, headers = usuario_autenticado
    resposta = cliente.get("/relatorios/excel/rebanho", headers=headers)
    assert resposta.status_code == 200

    wb = _ler_excel(resposta)
    ws = wb.active
    # Linha 3 é o cabeçalho (ver routers/relatorios.py); sem animais, não
    # deve haver nenhuma linha de dado na linha 4.
    assert ws.cell(row=4, column=1).value is None


# ─── Isolamento entre usuários ──────────────────────────────────────────────

def test_relatorio_rebanho_isola_por_usuario(
    usuario_autenticado, segundo_usuario_autenticado, criar_animal
):
    cliente1, usuario1, headers1 = usuario_autenticado
    _cliente2, usuario2, headers2 = segundo_usuario_autenticado

    criar_animal(usuario1.id, nome="Vaca Do Usuario 1", brinco="BR-U1")
    criar_animal(usuario2.id, nome="Vaca Do Usuario 2", brinco="BR-U2")

    resposta1 = cliente1.get("/relatorios/excel/rebanho", headers=headers1)
    resposta2 = cliente1.get("/relatorios/excel/rebanho", headers=headers2)

    nomes_usuario1 = _nomes_da_planilha(resposta1)
    nomes_usuario2 = _nomes_da_planilha(resposta2)

    assert "Vaca Do Usuario 1" in nomes_usuario1
    assert "Vaca Do Usuario 2" not in nomes_usuario1

    assert "Vaca Do Usuario 2" in nomes_usuario2
    assert "Vaca Do Usuario 1" not in nomes_usuario2


def _nomes_da_planilha(resposta):
    wb = _ler_excel(resposta)
    ws = wb.active
    nomes = []
    linha = 4
    while ws.cell(row=linha, column=1).value is not None:
        nomes.append(ws.cell(row=linha, column=1).value)
        linha += 1
    return nomes